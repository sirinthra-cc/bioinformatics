from sets import Set
import csv
import os
from openpyxl import Workbook


#file list
flist = ["../200_exomes/G2227-PJ.g.vcf", "../200_exomes/G2228-M.g.vcf"] # list of file name


def byChromosome(tup):
    if tup[0] == 'X':
        return 23 * 10000000000 + int(tup[1])
    elif tup[0] == 'Y':
        return 24 * 10000000000 + int(tup[1])
    elif tup[0] == 'M':
        return 0 + tup[1]
    elif len(tup[0][3:]) > 2: # A very strange case at the bottom of gcvf file
        return 100 * 10000000000 + int(tup[1])
    else:
        return int(tup[0]) * 10000000000 + int(tup[1])


def excelWrite(col,row,wdata):
    if col < 26:
        ws['%s%d' % (chr(col + 65), row)] = wdata
    else:
        ws['%s%s%d' % (chr(col/26+64), chr(col%26+65), row)] = wdata

db = "null"
Diff = Set()
cacheGene = Set()

# Find intersected differences
isFirstFile = True
print("Finding common variant...")

for file in flist:
    with open(file) as f:
        for line in f:
            if line[:11] == "#reference" and db == "null": db = line[11:]
            elif line[:11] == "#reference" and db != "null" and db != line[11:]:
                print("Warning, Unmatch reference detected. Found: " + line[11:] + ", Expected: " + db)
            elif line[:3] == "chr": # chromosome data
                s = line.split("\t")
                if s[4] != '<NON_REF>':
                    if isFirstFile:
                        Diff.add((s[0][3:], s[1]))
                    else:
                        cacheGene.add((s[0][3:], s[1]))
    if isFirstFile:
        isFirstFile = False
    else:
        Diff = Diff.intersection(cacheGene)
        cacheGene = Set()


Diff = list(Diff)
Diff = sorted(Diff, key=byChromosome)
## debug
## print(Diff)
## end of debug

print("Generating output")
# Generate list of differences from each file
#
# Writing Header

wb = Workbook()
ws = wb.active
ws['A1'] = "CHROM"
ws['B1'] = "POS"
ws['C1'] = "ID"
ws['D1'] = "REF"

fileNumber = 1
iterate = 0
for file in flist:
    with open(file) as f:
        filename = os.path.basename(f.name)
        filename = os.path.splitext(filename)[0]
        excelWrite(3 + fileNumber, 1, filename + ".GT")
        excelWrite(3 + fileNumber + len(flist), 1, filename + ".GT_1")
        excelWrite(3 + fileNumber + 2 * len(flist), 1, filename + ".DP")
        excelWrite(3 + fileNumber + 3 * len(flist), 1, filename + ".AD")
        for line in f:
            if line[:3] == "chr":
                s = line.split('\t')
                try:
                    if Diff[iterate][0] == s[0][3:] and Diff[iterate][1] == s[1]:
                        if fileNumber == 1: # first file should write data to first 4 column
                            ws['A%d' % (iterate + 2)] = "chr" + Diff[iterate][0]
                            ws['B%d' % (iterate + 2)] = int(Diff[iterate][1])
                            ws['C%d' % (iterate + 2)] = "."
                            ws['D%d' % (iterate + 2)] = s[3]
                        getFormat = s[8].split(':')
                        data = s[9].split(':')
                        excelWrite(3 + fileNumber, iterate + 2, s[3] + "/" + s[4][:-10])
                        try:
                            excelWrite(3 + fileNumber + len(flist), iterate + 2, data[getFormat.index("GT")])
                        except ValueError:
                            excelWrite(3 + fileNumber + len(flist), iterate + 2, "NaN")

                        try:
                            excelWrite(3 + fileNumber + len(flist) * 2, iterate + 2, int(data[getFormat.index("DP")]))
                        except ValueError:
                            excelWrite(3 + fileNumber + len(flist) * 2, iterate + 2, "NaN")
                        try:
                            excelWrite(3 + fileNumber + len(flist) * 3, iterate + 2, data[getFormat.index("AD")])
                        except ValueError:
                            excelWrite(3 + fileNumber + len(flist) * 3, iterate + 2, "NaN")
                        iterate = iterate + 1
                        if iterate >= len(Diff):
                            iterate = 0
                except IndexError:
                    print(f.name)
                    print(iterate)
                    print(s[0])
                    print(s[1])
    iterate = 0
    fileNumber = fileNumber + 1

wb.save("combined.xlsx")
